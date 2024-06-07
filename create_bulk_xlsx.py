#!/bin/env python

import argparse
import csv
import logging
import sys

from openpyxl import load_workbook


FIELD_NAMES = [
    'sender_name',
    'sender_phone',
    'sender_email',
    'package_value',
    'store_name',
    'store_id',
    'receiver_name',
    'receiver_phone',
    'receiver_email',
    'return_store_name',
    'return_store_id'
]


def get_input_reader(filename, fields, delimiter):
    if filename == '-':
        f = sys.stdin
    elif filename:
        f = open(filename, 'r', encoding='utf-8')
    else:
        f = open('/dev/null', 'r', encoding='utf-8')

    if fields:
        reader = csv.DictReader(f, fieldnames=fields, delimiter=delimiter)
    else:
        reader = csv.DictReader(f, delimiter=delimiter)

    return reader


def fill_row(row, field_names, default_values):
    output = []
    for field in field_names:
        if field in row:
            output.append(row[field])
        else:
            output.append(default_values[field])
    return output


def main():
    parser = argparse.ArgumentParser()

    parser.add_argument('--template', '-t',
                        help='The source xlsx file', required=True)
    parser.add_argument('--output', '-o',
                        help='The output xlsx file', required=True)
    parser.add_argument('--input', '-i',
                        help='The input file, or - for stdin (defaults to TSV)')
    parser.add_argument('--field', '-f', action='append',
                        help='A list of field names in the input')
    parser.add_argument('--delimiter', '-d', default='\t',
                        help='The delimiter of input file (defaults to tab)')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Verbosity level')

    # Generate CLI arguments programmatically
    for field in FIELD_NAMES:
        arg_name = '--' + field.replace('_', '-')
        help_str = 'Default value for ' + field.replace('_', ' ')
        parser.add_argument(arg_name, help=help_str)

    # Parse arguments
    args = parser.parse_args()

    # Set up logging
    logging_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=logging_level,
                        format='%(asctime)s [%(levelname)s] %(message)s')

    input_reader = get_input_reader(args.input, args.field, args.delimiter)
    default_values = {field: getattr(args, field) for field in FIELD_NAMES}

    # Create a list of rows, filling in missing fields from command-line args
    rows = []
    for row in input_reader:
        rows.append([None] + fill_row(row, FIELD_NAMES, default_values))
    if not rows:
        rows.append([None] + fill_row([], FIELD_NAMES, default_values))

    wb = load_workbook(filename=args.template)
    ws = wb.active
    row_id = 4
    for row in rows:
        logging.debug('Fill row: %s', row[1:])
        for cell_id, cell in enumerate(ws[row_id]):
            cell.value = row[cell_id]
        row_id += 1
    wb.save(args.output)


if __name__ == '__main__':
    main()
